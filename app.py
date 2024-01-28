from flask import Flask, request, jsonify, send_file, Blueprint, url_for
from flask_cors import CORS, cross_origin
from datetime import datetime, timedelta
import os
import json
import helper, env

app = Flask(__name__, template_folder="view", static_folder="lib")
app.config["JSON_SORT_KEYS"] = False
app.json.sort_keys = False
app.secret_key = env.app_secret_key
app.config['CORS_HEADERS'] = 'Content-Type'
cors = CORS(app, resources={r"/*": {"origins": "*"}})


route = "/export_excel"
@app.route(route, methods=['GET', 'POST'])
def export_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from io import BytesIO
    
    # json = {
    #     "filename" : "rekap absen siswa",
    #     "title": "Rekap absen siswa",
    #     "subtitle": "Periode: Januari 2024",
    #     "data": [
    #         {
    #             "id": 5,
    #             "token": "e389125964c81aa6e7a739bdc499c064",
    #             "tanggal": "2024-01-20 14:56:19",
    #             "siswa_nama": "Farauq Rifky"
    #         },
    #         {
    #             "id": 4,
    #             "token": "658f5738560fb2654280062115f0bb7d",
    #             "tanggal": "2024-01-20 07:54:52",
    #             "siswa_nama": "Farauq Rifky"
    #         },
    #         {
    #             "id": 3,
    #             "token": "7bfbd2a4cd40c84c6a03ba9b2b4c6d28",
    #             "tanggal": "2024-01-20 07:54:01",
    #             "siswa_nama": "Farauq Rifky"
    #         },
    #         {
    #             "id": 2,
    #             "token": "c18e3dd79e17a56bdbc5299cd1d66353",
    #             "tanggal": "2024-01-20 07:52:34",
    #             "siswa_nama": "Farauq Rifky"
    #         },
    #         {
    #             "id": 1,
    #             "token": "89a278cce3475b7f0d509ef3a9760145",
    #             "tanggal": "2024-01-18 00:00:00",
    #             "siswa_nama": "Aqillani"
    #         }
    #     ]
    # }
    
    json = request.get_json()
    print(json)
    filename = json["filename"]
    title = json["title"]
    subtitle = json["subtitle"]
    column_width = json["column_width"]
    cnotent = json["data"]

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Sheet1'

    firstData = cnotent[0]
    columnLength = len(firstData)
    columns = [list(item.keys()) for item in cnotent]
    columns = columns[0]
    data = [list(item.values()) for item in cnotent]
    
    columnStartIndex = 2 # B
    firstColumn = helper.num_to_col_excel(columnStartIndex)
    lastColumn = helper.num_to_col_excel(columnLength + 1) # 1 should be got from logic with columnStartIndex

    row_start = 1
    row_current = row_start

    row_current += 1
    display_judul = title
    sheet[f"{firstColumn}{row_current}"] = display_judul
    sheet[f"{firstColumn}{row_current}"].font = Font(bold=True, size=14)
    sheet[f"{firstColumn}{row_current}"].alignment = Alignment(horizontal='center', vertical='center')
    sheet.merge_cells(f"B{row_current}:{lastColumn}{row_current}")

    row_current += 1
    sheet[f"{firstColumn}{row_current}"] = subtitle
    sheet[f"{firstColumn}{row_current}"].font = Font(bold=True)

    row_current += 2
    for i, col in enumerate(columns):
        sheet[f"{helper.num_to_col_excel(columnStartIndex + i)}{row_current}"] = col
        sheet.column_dimensions[(f"{helper.num_to_col_excel(columnStartIndex + i)}")].width = column_width[i]

    for cell in helper.get_cells_in_range(sheet, f"{firstColumn}-{row_current}:{lastColumn}-{row_current}"):
        cell.font = Font(bold=True, size=14)
        cell.fill = PatternFill(start_color="c4c1c0", end_color="c4c1c0", fill_type="solid")
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    row_data_start = row_current + 1
    for i, record in enumerate(data):
        row_current += 1
        for j, col in enumerate(columns):
            sheet[f"{helper.num_to_col_excel(columnStartIndex + j)}{row_current}"] = data[i][j]
    row_data_end = row_current
    for cell in helper.get_cells_in_range(sheet, f"{firstColumn}-{row_data_start}:{lastColumn}-{row_data_end}"):
        cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    row_current += 2
    current_datetime = helper.get_local_time().strftime('%Y-%m-%d %H:%M:%S')
    sheet[f"{firstColumn}{row_current}"] = f"diunduh pada : {helper.tgl_indo(current_datetime, 'LONG')}"
    sheet[f"{firstColumn}{row_current}"].font = Font(italic=True)
    sheet.merge_cells(f"{firstColumn}{row_current}:{lastColumn}{row_current}")

    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    return send_file(excel_file, as_attachment=True, download_name=f'{filename}.xlsx', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
   

route = "/send_wa_1"
@app.route(route, methods=['POST'])
def send_wa_1():
    phone = request.form.get("phone")
    message = request.form.get("message")
    redirect = request.form.get("redirect")
    helper.send_wa_multipleSendText(phone, message, log=False)
    if not redirect == None:
        return redirect
    return helper.composeReply("SUCCESS", "Pesan dikirim")


route = "/toko_buah_webhook"
@app.route(route, methods=['POST'])
def toko_buah_webhook():
    
    return helper.composeReply("SUCCESS", "webhook processed from gatew")

 
if __name__ == '__main__':
    app.run(host = env.runHost, port = env.runPort, debug = env.runDebug)